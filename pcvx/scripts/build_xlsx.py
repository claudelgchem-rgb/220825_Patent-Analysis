#!/usr/bin/env python3
"""3번 산출물 — 특허 대장 xlsx. 시트 5종, 빈 셀 금지, Arial."""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import common  # noqa: E402

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)

LEDGER_COLUMNS = [
    ("id", "관리번호"), ("family_id", "패밀리 식별자"),
    ("representative_doc", "대표 문헌번호"), ("family_members", "패밀리 구성 문헌"),
    ("title_original", "제목(원문)"), ("title_ko", "제목(한국어)"),
    ("jurisdiction", "관할국"),
    ("application_number", "출원번호"), ("filing_date", "출원일"),
    ("publication_number", "공개번호"), ("publication_date", "공개일"),
    ("registration_number", "등록번호"), ("registration_date", "등록일"),
    ("registration_status", "등록 여부"), ("priority_date", "우선일"),
    ("expiry_estimate", "존속기간 만료 예정일"),
    ("legal_status", "법적 상태"), ("legal_status_checked_on", "법적상태 확인일"),
    ("applicant", "출원인"), ("current_assignee", "현재 권리자"),
    ("assignment_history", "양도 이력"), ("inventors", "발명자"),
    ("ipc", "국제특허분류"), ("cpc", "협력적 특허분류"),
    ("source_url", "원문 URL"),
]


def style_sheet(ws, widths: list[int]) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(ws.max_row, 1)}"
    ws.row_dimensions[1].height = 30


def sheet_ledger(wb, records: list, grades: dict) -> None:
    ws = wb.active
    ws.title = "특허대장"
    ws.append([label for _, label in LEDGER_COLUMNS] + ["항목 신뢰도 요약"])
    for rec in records:
        rid = rec.get("id", "?")
        row = [common.cell(rec.get(key)) for key, _ in LEDGER_COLUMNS]
        marks = [f"{f}={g}" for (r, f), g in grades.items() if r == rid]
        row.append(common.cell(", ".join(sorted(marks))))
        ws.append(row)
    style_sheet(ws, [10, 20, 20, 26, 40, 40, 8] + [16] * 11 + [22, 22, 24, 24, 20, 20, 38, 46])


def sheet_claims(wb, records: list) -> None:
    ws = wb.create_sheet("청구항")
    ws.append(["관리번호", "대표 문헌번호", "독립항 원문", "구성요소", "요소의 역할",
               "권리범위에 미치는 영향", "종속항 요약", "초보자용 평문 해설"])
    for rec in records:
        rid = common.cell(rec.get("id"))
        doc = common.cell(rec.get("representative_doc"))
        claim = common.cell(rec.get("independent_claim_text"))
        dep = common.cell(rec.get("dependent_claim_summary"))
        plain = common.cell(rec.get("plain_explanation"))
        elements = rec.get("claim_elements") or [
            {"element": common.FAILURE_MARK, "role": common.FAILURE_MARK,
             "scope_effect": common.FAILURE_MARK}
        ]
        for pos, el in enumerate(elements):
            ws.append([
                rid, doc,
                claim if pos == 0 else "(위 셀과 같은 특허)",
                common.cell(el.get("element")),
                common.cell(el.get("role")),
                common.cell(el.get("scope_effect")),
                dep if pos == 0 else "(위 셀과 같은 특허)",
                plain if pos == 0 else "(위 셀과 같은 특허)",
            ])
    style_sheet(ws, [10, 22, 60, 26, 34, 40, 40, 50])


def sheet_grades(wb, verification: dict) -> None:
    ws = wb.create_sheet("신뢰도")
    ws.append(["항목 식별자", "관리번호", "항목", "등급", "판정 근거", "출처",
               "재조회 확인일", "정정 여부", "정정 전 값", "정정 후 값"])
    for item in verification.get("items", []):
        cid = common.cell(item.get("claim_id"))
        ws.append([
            cid,
            common.cell(cid.split(".")[0]),
            common.cell(cid.split(".")[-1]),
            common.cell(item.get("grade")),
            common.cell(item.get("basis")),
            common.cell(item.get("sources")),
            common.cell(item.get("checked_on")),
            "정정함" if item.get("corrected") else "정정 없음",
            common.cell(item.get("original_value"), "해당 없음"),
            common.cell(item.get("corrected_value"), "해당 없음"),
        ])
    style_sheet(ws, [26, 12, 22, 8, 50, 46, 16, 12, 28, 28])


def sheet_queries(wb, queries: dict) -> None:
    ws = wb.create_sheet("검색식")
    ws.append(["식 번호", "유형", "검색식", "의도"])
    for q in queries.get("queries", []):
        ws.append([common.cell(q.get("id")), common.cell(q.get("type")),
                   common.cell(q.get("expression")), common.cell(q.get("intent"))])
    ws.append(["—", "분류코드", "—", "아래는 국제특허분류·협력적 특허분류 후보다"])
    for c in queries.get("classifications", []):
        ws.append(["—", common.cell(c.get("system"), "분류"),
                   common.cell(c.get("code")), common.cell(c.get("note"))])
    style_sheet(ws, [10, 12, 60, 60])


def sheet_sources(wb, attempts: dict, records: list) -> None:
    ws = wb.create_sheet("출처")
    ws.append(["구분", "소스", "URL", "시간 창", "수집 건수", "접속일", "비고"])
    today = common.today()
    for a in attempts.get("attempts", []):
        ws.append([
            "검색 시도", common.cell(a.get("source")), common.cell(a.get("url")),
            common.cell(a.get("window")), common.cell(a.get("hits"), "0"),
            common.cell(a.get("accessed_on"), today),
            common.cell(a.get("failure") or a.get("note"), "정상 수행"),
        ])
    for rec in records:
        ws.append([
            "문헌 원문", common.cell(rec.get("jurisdiction")),
            common.cell(rec.get("source_url")), "해당 없음", "1",
            common.cell(rec.get("legal_status_checked_on"), today),
            common.cell(rec.get("representative_doc")),
        ])
    style_sheet(ws, [12, 10, 60, 14, 10, 14, 30])


def main() -> int:
    env = common.env()
    records = (common.read_json(common.WORKSPACE / "records.json", {}) or {}).get("records", [])
    verification = common.read_json(common.WORKSPACE / "a7_verification.json", {}) or {}
    queries = common.read_json(common.WORKSPACE / "a1_queries.json", {}) or {}
    attempts = common.read_json(common.WORKSPACE / "source_attempts.json", {}) or {}

    if not records:
        print("[실패] records.json 에 레코드가 없다. A3~A5 를 먼저 완료한다.")
        return 1

    grades = {
        (i["claim_id"].split(".")[0], i["claim_id"].split(".")[-1]): i.get("grade", "미판정")
        for i in verification.get("items", []) if i.get("claim_id")
    }

    wb = Workbook()
    sheet_ledger(wb, records, grades)
    sheet_claims(wb, records)
    sheet_grades(wb, verification)
    sheet_queries(wb, queries)
    sheet_sources(wb, attempts, records)

    out = common.output_dir() / f"PCVX_특허대장_{env['as_of_compact']}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[기준일: {env['as_of']} / 도구: build_xlsx]")
    print(f"-> {out}  ({out.stat().st_size:,} bytes, 시트 {len(wb.sheetnames)}개: {', '.join(wb.sheetnames)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
